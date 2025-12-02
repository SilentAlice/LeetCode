#!/usr/bin/env python3
"""
Excel Seat Checker
Reads two Excel files:
1. Name list - contains people names
2. Seat graph - contains seat numbers and people names across multiple sheets

Checks:
- People in name list who don't have seats
- Seats that don't have corresponding people in the name list
"""

import pandas as pd
import sys
from pathlib import Path
from typing import Dict, Set, List, Tuple, Optional
import argparse
import re
import openpyxl


def normalize_name(name):
    """Normalize name for comparison (strip whitespace, convert to lowercase)"""
    if pd.isna(name) or name is None:
        return None
    return str(name).strip().lower()


def contains_chinese(text):
    """Check if text contains Chinese characters"""
    if not text or pd.isna(text):
        return False
    text_str = str(text)
    # Check for Chinese characters (CJK Unified Ideographs)
    return bool(re.search(r'[\u4e00-\u9fff]', text_str))


def is_seat_number(text):
    """
    Check if text matches seat number format.
    Formats: "1-001S" (floor-seat) or "加座1" (additional seat)
    """
    if not text or pd.isna(text):
        return False
    text_str = str(text).strip()
    
    # Pattern 1: floor-seat format like "1-001S", "2-123S", etc.
    pattern1 = re.match(r'^\d+-\d+S$', text_str)
    
    # Pattern 2: additional seat format like "加座1", "加座2", etc.
    pattern2 = re.match(r'^加座\d+$', text_str)
    
    return bool(pattern1 or pattern2)


def is_merged_cell(worksheet, row, col):
    """Check if a cell is part of a merged cell range"""
    cell = worksheet.cell(row=row, column=col)
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False


def read_name_list(file_path: str, name_column: str = None) -> Set[str]:
    """
    Read name list from Excel file and return set of normalized names.
    
    Args:
        file_path: Path to the name list Excel file
        name_column: Column name containing names. If None, uses first column.
    
    Returns:
        Set of normalized names
    """
    try:
        # Read all sheets and combine names
        excel_file = pd.ExcelFile(file_path)
        all_names = set()
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            if df.empty:
                continue
            
            # Use specified column or first column
            if name_column and name_column in df.columns:
                col = name_column
            else:
                col = df.columns[0]
            
            # Extract names
            names = df[col].dropna()
            normalized_names = {normalize_name(name) for name in names if normalize_name(name)}
            all_names.update(normalized_names)
            
            print(f"  Sheet '{sheet_name}': Found {len(normalized_names)} names")
        
        return all_names
    
    except Exception as e:
        print(f"Error reading name list file: {e}")
        sys.exit(1)


def read_name_list_full(file_path: str, name_column: str = None) -> pd.DataFrame:
    """
    Read full name list from Excel file with all columns.
    
    Args:
        file_path: Path to the name list Excel file
        name_column: Column name containing names. If None, uses first column.
    
    Returns:
        DataFrame with all columns from name list
    """
    try:
        # Read all sheets and combine into one DataFrame
        excel_file = pd.ExcelFile(file_path)
        all_dfs = []
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            if df.empty:
                continue
            
            all_dfs.append(df)
        
        if not all_dfs:
            return pd.DataFrame()
        
        # Combine all sheets
        combined_df = pd.concat(all_dfs, ignore_index=True)
        return combined_df
    
    except Exception as e:
        print(f"Error reading name list file: {e}")
        sys.exit(1)


def read_seat_graph(file_path: str, max_rows: int = 1000, max_cols: int = 200) -> Dict[str, Dict[str, str]]:
    """
    Read seat graph from Excel file with multiple sheets.
    Each sheet represents a building. Seats are identified by:
    - Finding cells with Chinese characters (names)
    - Checking the cell above for seat numbers (format: "1-001S" or "加座1")
    - Ignoring merged cells
    
    Args:
        file_path: Path to the seat graph Excel file
        max_rows: Maximum number of rows to scan (default: 1000)
        max_cols: Maximum number of columns to scan (default: 200)
    
    Returns:
        Dictionary: {building_number: {seat_num: person_name}}
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        seat_data = {}
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            building_seats = {}
            
            # Scan through cells (skip first row as seat numbers are above names)
            found_seats = 0
            for row in range(2, min(worksheet.max_row + 1, max_rows + 1)):
                for col in range(1, min(worksheet.max_column + 1, max_cols + 1)):
                    # Skip if this cell is part of a merged cell
                    if is_merged_cell(worksheet, row, col):
                        continue
                    
                    # Get current cell value (potential name)
                    current_cell = worksheet.cell(row=row, column=col)
                    current_value = current_cell.value
                    
                    # Check if current cell contains Chinese characters (likely a name)
                    if current_value and contains_chinese(str(current_value)):
                        # Check the cell above for seat number
                        top_cell = worksheet.cell(row=row - 1, column=col)
                        top_value = top_cell.value
                        
                        if top_value and is_seat_number(str(top_value)):
                            seat_num = str(top_value).strip()
                            person_name = str(current_value).strip()
                            person_name_normalized = normalize_name(person_name)
                            
                            # Store the seat assignment
                            building_seats[seat_num] = person_name_normalized
                            found_seats += 1
            
            seat_data[sheet_name] = building_seats
            print(f"  Building '{sheet_name}': Found {found_seats} seat assignments")
        
        workbook.close()
        return seat_data
    
    except Exception as e:
        print(f"Error reading seat graph file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def check_seats(name_list: Set[str], seat_graph: Dict[str, Dict[str, str]]) -> Tuple[List[str], List[Tuple[str, str, str]]]:
    """
    Check for missing seats and empty seats.
    
    Args:
        name_list: Set of normalized names from name list
        seat_graph: Dictionary of {building: {seat_num: person_name}}
    
    Returns:
        Tuple of (people_without_seats, empty_seats)
        empty_seats is list of (building, seat_num, person_name) tuples
    """
    # Collect all names that have seats
    names_with_seats = set()
    all_seats = []
    
    for building, seats in seat_graph.items():
        for seat_num, person_name in seats.items():
            all_seats.append((building, seat_num, person_name))
            if person_name:
                names_with_seats.add(person_name)
    
    # Find people in name list without seats
    people_without_seats = list(name_list - names_with_seats)
    
    # Find seats without corresponding people in name list
    empty_seats = []
    for building, seat_num, person_name in all_seats:
        if not person_name or person_name not in name_list:
            empty_seats.append((building, seat_num, person_name))
    
    return people_without_seats, empty_seats


def create_output_dataframe(name_list_df: pd.DataFrame, name_column: str, 
                           seat_graph: Dict[str, Dict[str, str]], 
                           name_list_set: Set[str]) -> pd.DataFrame:
    """
    Create output DataFrame with name list info and seat assignments.
    
    Args:
        name_list_df: Full DataFrame from name list file
        name_column: Column name containing names
        seat_graph: Dictionary of {building: {seat_num: person_name}}
        name_list_set: Set of normalized names from name list
    
    Returns:
        DataFrame with all name list columns plus 'Seat' column, plus rows for unmatched seats
    """
    # Create reverse mapping: normalized_name -> (building, seat_num)
    name_to_seat = {}
    for building, seats in seat_graph.items():
        for seat_num, person_name in seats.items():
            if person_name and person_name in name_list_set:
                if person_name not in name_to_seat:
                    name_to_seat[person_name] = []
                name_to_seat[person_name].append(f"{building}-{seat_num}")
    
    # Determine name column
    if name_column and name_column in name_list_df.columns:
        name_col = name_column
    else:
        name_col = name_list_df.columns[0]
    
    # Create a copy of the name list DataFrame
    output_df = name_list_df.copy()
    
    # Add Seat column
    output_df['Seat'] = ''
    
    # Fill in seat information for people who have seats
    for idx, row in output_df.iterrows():
        name = row[name_col]
        if pd.notna(name):
            normalized_name = normalize_name(name)
            if normalized_name in name_to_seat:
                # Join multiple seats if person has multiple seats
                output_df.at[idx, 'Seat'] = ', '.join(name_to_seat[normalized_name])
    
    # Find unmatched seats (empty or with people not in name list)
    unmatched_seats = []
    for building, seats in seat_graph.items():
        for seat_num, person_name in seats.items():
            if not person_name or person_name not in name_list_set:
                unmatched_seats.append({
                    'building': building,
                    'seat_num': seat_num,
                    'person_name': person_name if person_name else '(empty)'
                })
    
    # Add rows for unmatched seats
    if unmatched_seats:
        # Create DataFrame for unmatched seats
        unmatched_rows = []
        for seat_info in unmatched_seats:
            row_data = {}
            # Fill all original columns with empty values
            for col in output_df.columns:
                if col == 'Seat':
                    row_data[col] = f"{seat_info['building']}-{seat_info['seat_num']}"
                else:
                    row_data[col] = ''
            unmatched_rows.append(row_data)
        
        unmatched_df = pd.DataFrame(unmatched_rows)
        # Ensure columns match exactly
        unmatched_df = unmatched_df[output_df.columns]
        # Combine with original DataFrame
        output_df = pd.concat([output_df, unmatched_df], ignore_index=True)
    
    return output_df


def main():
    parser = argparse.ArgumentParser(
        description='Check seat assignments between name list and seat graph Excel files'
    )
    parser.add_argument('name_list_file', help='Path to name list Excel file')
    parser.add_argument('seat_graph_file', help='Path to seat graph Excel file')
    parser.add_argument('--name-column', help='Column name for names in name list (auto-detect if not specified)')
    parser.add_argument('--max-rows', type=int, default=1000, help='Maximum rows to scan in seat graph (default: 1000)')
    parser.add_argument('--max-cols', type=int, default=200, help='Maximum columns to scan in seat graph (default: 200)')
    parser.add_argument('--output', help='Output file path for results (Excel format, default: seat_assignment_output.xlsx)')
    
    args = parser.parse_args()
    
    # Check if files exist
    if not Path(args.name_list_file).exists():
        print(f"Error: Name list file not found: {args.name_list_file}")
        sys.exit(1)
    
    if not Path(args.seat_graph_file).exists():
        print(f"Error: Seat graph file not found: {args.seat_graph_file}")
        sys.exit(1)
    
    print("=" * 60)
    print("Excel Seat Checker")
    print("=" * 60)
    
    # Read name list (both for checking and full data)
    print(f"\nReading name list from: {args.name_list_file}")
    name_list = read_name_list(args.name_list_file, args.name_column)
    print(f"Total unique names in name list: {len(name_list)}")
    
    # Read full name list DataFrame
    print("Reading full name list data...")
    name_list_df = read_name_list_full(args.name_list_file, args.name_column)
    print(f"Total rows in name list: {len(name_list_df)}")
    
    # Read seat graph
    print(f"\nReading seat graph from: {args.seat_graph_file}")
    print(f"Scanning up to {args.max_rows} rows and {args.max_cols} columns per sheet")
    seat_graph = read_seat_graph(args.seat_graph_file, args.max_rows, args.max_cols)
    total_seats = sum(len(seats) for seats in seat_graph.values())
    print(f"Total seats across all buildings: {total_seats}")
    
    # Create output DataFrame
    print("\nCreating output DataFrame...")
    output_df = create_output_dataframe(name_list_df, args.name_column, seat_graph, name_list)
    
    # Perform checks for reporting
    print("\n" + "=" * 60)
    print("Analysis Results")
    print("=" * 60)
    
    people_without_seats, empty_seats = check_seats(name_list, seat_graph)
    
    # Report people without seats
    print(f"\n1. People in name list WITHOUT seats: {len(people_without_seats)}")
    if people_without_seats:
        for name in sorted(people_without_seats):
            print(f"   - {name}")
    else:
        print("   ✓ All people in name list have seats!")
    
    # Report empty seats
    print(f"\n2. Seats WITHOUT corresponding people in name list: {len(empty_seats)}")
    if empty_seats:
        for building, seat_num, person_name in sorted(empty_seats):
            person_display = person_name if person_name else "(empty)"
            print(f"   - Building '{building}', Seat '{seat_num}': {person_display}")
    else:
        print("   ✓ All seats have corresponding people in name list!")
    
    # Summary statistics
    names_with_seats = name_list - set(people_without_seats)
    print(f"\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"Total names in list: {len(name_list)}")
    print(f"Names with seats: {len(names_with_seats)}")
    print(f"Names without seats: {len(people_without_seats)}")
    print(f"Total seats: {total_seats}")
    print(f"Empty/unassigned seats: {len(empty_seats)}")
    print(f"Total output rows: {len(output_df)}")
    
    # Save to output file
    output_file = args.output if args.output else "seat_assignment_output.xlsx"
    print(f"\nSaving results to: {output_file}")
    output_df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"✓ Results saved successfully!")


if __name__ == '__main__':
    main()
